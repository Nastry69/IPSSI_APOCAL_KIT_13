"""
Construction de l'export RGPD (droit à la portabilité, art. 20).

On assemble un dictionnaire JSON-sérialisable regroupant TOUTES les données
personnelles d'un utilisateur : son compte, son profil (dont le consentement),
ses quiz et les questions associées. Aucune donnée d'un autre utilisateur n'est
incluse (isolation stricte par `user`).

Ce module fournit aussi le RENDU de cet export en plusieurs formats de
téléchargement direct : JSON, CSV (tabulaire), HTML (page lisible) et XLSX
(classeur Excel). Chaque fonction `render_*` part du dict `build_user_export`
et renvoie des `bytes` prêts à être servis en pièce jointe.
"""

import csv
import io
import json

from django.utils.html import escape

from .models import get_or_create_profile


def build_user_export(user) -> dict:
    """Renvoie un dict JSON-sérialisable de toutes les données de `user`.

    Ne fait AUCUN accès aux données d'autres utilisateurs : on part de `user`
    et on suit ses relations (`user.quizzes` -> `quiz.questions`).
    """
    profile = get_or_create_profile(user)

    quizzes = []
    # `related_name="quizzes"` sur Quiz.user ; `related_name="questions"` sur Question.quiz
    for quiz in user.quizzes.all().order_by("created_at"):
        questions = [
            {
                "index": q.index,
                "prompt": q.prompt,
                "options": q.options,
                "correct_index": q.correct_index,
                "selected_index": q.selected_index,
            }
            for q in quiz.questions.all().order_by("index")
        ]
        quizzes.append(
            {
                "title": quiz.title,
                "source_text": quiz.source_text,
                "score": quiz.score,
                "created_at": quiz.created_at.isoformat(),
                "updated_at": quiz.updated_at.isoformat(),
                "questions": questions,
            }
        )

    return {
        "account": {
            "id": user.id,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "date_joined": user.date_joined.isoformat(),
        },
        "profile": {
            "email_verified": profile.email_verified,
            "created_at": profile.created_at.isoformat(),
            "consentement": {
                "consent_accepted_at": (
                    profile.consent_accepted_at.isoformat() if profile.consent_accepted_at else None
                ),
                "consent_version": profile.consent_version or None,
            },
        },
        "quizzes": quizzes,
    }


# ---------------------------------------------------------------------------
# Rendu multi-format de l'export (téléchargement direct)
#
# Chaque fonction part du dict `build_user_export(user)` et renvoie des `bytes`.
# On garde un CONTENU cohérent d'un format à l'autre (compte, profil, quiz,
# questions), simplement présenté différemment selon le format.
# ---------------------------------------------------------------------------

# Formats supportés -> extension de fichier + type MIME.
EXPORT_FORMATS = {
    "json": ("json", "application/json; charset=utf-8"),
    "csv": ("csv", "text/csv; charset=utf-8"),
    "html": ("html", "text/html; charset=utf-8"),
    "xlsx": (
        "xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
}


def _account_rows(data: dict) -> list[tuple[str, str]]:
    """Aplati la section compte en lignes (clé lisible, valeur)."""
    account = data["account"]
    return [
        ("ID", account["id"]),
        ("Email", account["email"]),
        ("Prénom", account["first_name"]),
        ("Nom", account["last_name"]),
        ("Inscription", account["date_joined"]),
    ]


def _profile_rows(data: dict) -> list[tuple[str, str]]:
    """Aplati la section profil (dont le consentement) en lignes."""
    profile = data["profile"]
    consent = profile["consentement"]
    return [
        ("Email vérifié", profile["email_verified"]),
        ("Profil créé le", profile["created_at"]),
        ("Consentement accepté le", consent["consent_accepted_at"]),
        ("Version du consentement", consent["consent_version"]),
    ]


# En-têtes des colonnes pour la partie tabulaire (quiz + questions).
_QUIZ_COLUMNS = ["title", "source_text", "score", "created_at", "updated_at"]
_QUESTION_COLUMNS = [
    "quiz_title",
    "index",
    "prompt",
    "options",
    "correct_index",
    "selected_index",
]


def render_json(data: dict) -> bytes:
    """Rendu JSON : le dict d'export sérialisé (indenté, UTF-8, accents préservés)."""
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")


def render_csv(data: dict) -> bytes:
    """Rendu CSV TABULAIRE : sections compte/profil, puis une ligne par quiz et
    une ligne par question, avec des en-têtes de section lisibles."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Section compte
    writer.writerow(["# Compte"])
    writer.writerow(["Champ", "Valeur"])
    for label, value in _account_rows(data):
        writer.writerow([label, value])
    writer.writerow([])

    # Section profil
    writer.writerow(["# Profil"])
    writer.writerow(["Champ", "Valeur"])
    for label, value in _profile_rows(data):
        writer.writerow([label, value])
    writer.writerow([])

    # Section quiz (une ligne par quiz)
    writer.writerow(["# Quiz"])
    writer.writerow(_QUIZ_COLUMNS)
    for quiz in data["quizzes"]:
        writer.writerow([quiz.get(col, "") for col in _QUIZ_COLUMNS])
    writer.writerow([])

    # Section questions (une ligne par question, rattachée à son quiz)
    writer.writerow(["# Questions"])
    writer.writerow(_QUESTION_COLUMNS)
    for quiz in data["quizzes"]:
        for question in quiz["questions"]:
            writer.writerow(
                [
                    quiz["title"],
                    question["index"],
                    question["prompt"],
                    " | ".join(str(opt) for opt in question["options"]),
                    question["correct_index"],
                    question["selected_index"],
                ]
            )

    # BOM UTF-8 pour qu'Excel ouvre correctement les accents.
    return buffer.getvalue().encode("utf-8-sig")


def _html_table(rows: list[tuple[str, str]]) -> str:
    """Construit un tableau HTML clé/valeur (contenu échappé)."""
    lines = ["<table>"]
    for label, value in rows:
        lines.append(f"<tr><th>{escape(str(label))}</th><td>{escape(str(value))}</td></tr>")
    lines.append("</table>")
    return "\n".join(lines)


def render_html(data: dict) -> bytes:
    """Rendu HTML : une page simple et lisible (titres + tableaux).

    IMPORTANT : tout le contenu est échappé via `django.utils.html.escape` — les
    champs `prompt`, `source_text`, `title`… proviennent de l'utilisateur et
    pourraient contenir du HTML/JS.
    """
    parts = [
        "<!DOCTYPE html>",
        '<html lang="fr">',
        "<head>",
        '<meta charset="utf-8">',
        "<title>Mes données EduTutor IA</title>",
        "<style>"
        "body{font-family:sans-serif;max-width:900px;margin:2rem auto;padding:0 1rem;}"
        "table{border-collapse:collapse;width:100%;margin-bottom:1.5rem;}"
        "th,td{border:1px solid #ccc;padding:.4rem .6rem;text-align:left;vertical-align:top;}"
        "th{background:#f4f4f4;}"
        "h1{margin-bottom:.2rem;}h2{margin-top:2rem;}"
        "</style>",
        "</head>",
        "<body>",
        "<h1>Mes données personnelles — EduTutor IA</h1>",
        "<p>Export RGPD (droit à la portabilité, art. 20).</p>",
        "<h2>Compte</h2>",
        _html_table(_account_rows(data)),
        "<h2>Profil</h2>",
        _html_table(_profile_rows(data)),
        "<h2>Quiz</h2>",
    ]

    quizzes = data["quizzes"]
    if not quizzes:
        parts.append("<p>Aucun quiz.</p>")
    for quiz in quizzes:
        parts.append(f"<h3>{escape(str(quiz['title']))}</h3>")
        parts.append(
            _html_table(
                [
                    ("Texte source", quiz["source_text"]),
                    ("Score", quiz["score"]),
                    ("Créé le", quiz["created_at"]),
                    ("Mis à jour le", quiz["updated_at"]),
                ]
            )
        )
        if quiz["questions"]:
            rows = ["<table>"]
            rows.append(
                "<tr><th>#</th><th>Question</th><th>Options</th>"
                "<th>Bonne réponse</th><th>Réponse donnée</th></tr>"
            )
            for question in quiz["questions"]:
                options = " | ".join(str(opt) for opt in question["options"])
                rows.append(
                    "<tr>"
                    f"<td>{escape(str(question['index']))}</td>"
                    f"<td>{escape(str(question['prompt']))}</td>"
                    f"<td>{escape(options)}</td>"
                    f"<td>{escape(str(question['correct_index']))}</td>"
                    f"<td>{escape(str(question['selected_index']))}</td>"
                    "</tr>"
                )
            rows.append("</table>")
            parts.append("\n".join(rows))
        else:
            parts.append("<p>Aucune question.</p>")

    parts.extend(["</body>", "</html>"])
    return "\n".join(parts).encode("utf-8")


def render_xlsx(data: dict) -> bytes:
    """Rendu XLSX : un classeur Excel via openpyxl, avec trois feuilles
    (Compte/Profil, Quiz, Questions)."""
    # Import local : openpyxl n'est requis QUE pour ce format (dépendance
    # présente en conteneur). On évite de la charger pour les autres formats.
    from openpyxl import Workbook

    workbook = Workbook()

    # Feuille 1 : Compte + Profil (paires champ/valeur).
    sheet = workbook.active
    sheet.title = "Compte-Profil"
    sheet.append(["Section", "Champ", "Valeur"])
    for label, value in _account_rows(data):
        sheet.append(["Compte", label, value])
    for label, value in _profile_rows(data):
        sheet.append(["Profil", label, value])

    # Feuille 2 : Quiz (une ligne par quiz).
    quiz_sheet = workbook.create_sheet("Quiz")
    quiz_sheet.append(_QUIZ_COLUMNS)
    for quiz in data["quizzes"]:
        quiz_sheet.append([quiz.get(col, "") for col in _QUIZ_COLUMNS])

    # Feuille 3 : Questions (une ligne par question).
    question_sheet = workbook.create_sheet("Questions")
    question_sheet.append(_QUESTION_COLUMNS)
    for quiz in data["quizzes"]:
        for question in quiz["questions"]:
            question_sheet.append(
                [
                    quiz["title"],
                    question["index"],
                    question["prompt"],
                    " | ".join(str(opt) for opt in question["options"]),
                    question["correct_index"],
                    question["selected_index"],
                ]
            )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# Table de dispatch format -> fonction de rendu.
RENDERERS = {
    "json": render_json,
    "csv": render_csv,
    "html": render_html,
    "xlsx": render_xlsx,
}


def render_export(user, fmt: str) -> tuple[bytes, str, str]:
    """Rend l'export de `user` au format `fmt`.

    Renvoie (contenu en bytes, type MIME, extension de fichier).
    Lève `ValueError` si le format n'est pas supporté.
    """
    if fmt not in EXPORT_FORMATS:
        raise ValueError(f"Format d'export non supporté : {fmt!r}")
    extension, content_type = EXPORT_FORMATS[fmt]
    data = build_user_export(user)
    content = RENDERERS[fmt](data)
    return content, content_type, extension
